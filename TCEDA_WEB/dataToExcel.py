from dataFromDB import *
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.chart import BarChart, ScatterChart, Reference, Series
from django.http import HttpResponse

"""
For all the charts and just the data in general--need to figure out how to arrange the data in the excel file based
on need- charts-figure out how to best get the chart needed with the data
"""


def race_toExcel():
    """
    generates excel doc with race data
    :return: HttpResponse(excel file)
    """
    workbook = Workbook()
    worksheet = workbook.active
    race_data = population_by_raceDB()
    worksheet.append(race_data['header'])
    percentage_list = []
    for race, data in race_data['data'].iteritems():
        worksheet.append(data.values())
    worksheet.append(percentage_list)
    worksheet.append([race_data['source']])
    response = HttpResponse(content=save_virtual_workbook(workbook),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=race.xlsx'
    return response


def income_dist_toExcel():
    """
    generates excel doc with income data
    :return: HttpResponse(excel file)
    """
    workbook = Workbook()
    worksheet = workbook.active
    house_data = house_income_distDB()
    worksheet.append(house_data['header'])
    percentage_list = []
    for race, data in house_data['data'].iteritems():
        worksheet.append(data.values())
    worksheet.append(percentage_list)
    worksheet.append([house_data['source']])
    response = HttpResponse(content=save_virtual_workbook(workbook),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=income_dist.xlsx'
    return response


def language_dist_toExcel():
    """
    generates excel doc with language data
    :return: HttpResponse(excel file)
    """
    workbook = Workbook()
    worksheet = workbook.active
    language_data = language_distDB()
    worksheet.append(language_data['header'])
    for year, data in language_data['data'].iteritems():
        for x, y in data.iteritems():
            worksheet.append(y)
    worksheet.append([language_data['source']])
    response = HttpResponse(content=save_virtual_workbook(workbook),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=language.xlsx'
    return response


def householdIncome_toExcel():
    """
    generates excel doc and chart with income data
    :return: HttpResponse(excel file)
    """
    # just add more-- locations and change up code a bit
    workbook = Workbook()
    worksheet = workbook.active
    household_data = household_incomeDB()
    worksheet.append(list(household_data['header'].values()))
    for year, data in household_data['data'].iteritems():
        for x, y in data.iteritems():
            worksheet.append(y.values())
    """
    years = []
    for year, data in household_data['data'].iteritems():
        years.append(year)

    location_dict = {}
    for year, data in household_data['data'].iteritems():
        # print year

        for key, location in data.iteritems():
            if location['location'] in location_dict:
                location_dict[location['location']].append({
                    'year': year,
                    'income': location['income']
                })
            else:
                location_dict[location['location']] = [{
                    'year': year,
                    'income': location['income']
                }]
    """  # plz figure out how to make cluster bar chart :)
    worksheet.append([household_data['source']])
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "Bar Chart"
    chart1.y_axis.title = household_data['header']['income']
    chart1.x_axis.title = household_data['header']['location']

    data = Reference(worksheet, min_col=3, min_row=1, max_row=((len(location_dict) + 1) * 2) - 1,
                     max_col=3)
    cats = Reference(worksheet, min_col=1, min_row=2, max_row=(len(location_dict) + 1) * 2,
                     max_col=2)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.shape = 4
    worksheet.add_chart(chart1)

    response = HttpResponse(content=save_virtual_workbook(workbook),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=household.xlsx'
    return response


def population_toExcel():
    """
    generates excel doc and chart with population data
    :return: HttpResponse(excel file)
    """
    workbook = Workbook()
    worksheet = workbook.active
    population_data = population_from_db()
    worksheet.append(list(population_data['header'].values()))
    for date, data in population_data['data'].iteritems():
        worksheet.append(data.values())
    worksheet.append([population_data['source']])
    chart = ScatterChart()
    chart.scatterStyle = 'marker'
    chart.title = "Tuolumne County Population"
    chart.style = 13
    chart.x_axis.title = population_data['header']['year']
    chart.y_axis.title = population_data['header']['population']

    xvalues = Reference(worksheet, min_col=3, min_row=1, max_row=len(population_data['data']) + 1)
    values = Reference(worksheet, min_col=1, min_row=2, max_row=len(population_data['data']) + 1)
    series = Series(xvalues, values, title_from_data=True)
    chart.series.append(series)

    worksheet.add_chart(chart)
    response = HttpResponse(content=save_virtual_workbook(workbook),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=population.xlsx'

    return response
